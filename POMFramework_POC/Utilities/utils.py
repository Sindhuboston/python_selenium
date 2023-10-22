import openpyxl
import inspect
import logging
from openpyxl import Workbook, load_workbook
from Config.config import TestData


# The 'Utils' class serves as a utility class within a test automation framework and provides various helper methods.
# It is designed to offer convenient functions for logging, reading data from Excel files, and other utility functions.
# The key methods and features within this class include:
# - 'log_to_file_output': A method to create and configure logging for test activities, which can be used for debugging and reporting.
# - 'get_data': A method for reading data from an Excel file, specified by the 'TestData' class, based on the provided column name.
# Overall, the 'Utils' class simplifies common tasks and enhances test automation by providing reusable utility methods.

class Utils:
    @staticmethod
    def log_to_file_output(logLevel=logging.DEBUG):
        # 1. Set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        # 1. create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        # 2. create console handler or file handler and set the log level
        fh = logging.FileHandler("automation.log")
        # 3. create formatter - how you want your logs to be formatted
        formatter1 = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s - %(message)s',
                                       datefmt='%m/%d/%Y %I:%M:%S %p')
        # 4. add formatter to console or file handler
        fh.setFormatter(formatter1)
        # 5. add console handler to logger
        logger.addHandler(fh)
        return logger

    # A method for reading data from an Excel file, specified by the 'TestData' class, based on the
    # provided column name. The 'get_data' method retrieves data from an Excel file, allowing easy access to test
    # data during test case execution. It will search for the test case name and return the data based on the
    # specified column. If duplicate testcases are found 1st one will be executed.
    @staticmethod
    def get_data(column_name):
        # Specify the Excel file path
        excel_file_path = TestData.EXCEL_FILE_PATH
        sheet_name = TestData.SHEET_NAME
        test_case_name = TestData.TESTCASE_NAME
        # print(f"test_case_name ={test_case_name}")

        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            worksheet = workbook[sheet_name]
            headings = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
            row_data = {}

            # Find the row number by test case name
            row_number = None
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
                for cell in row:
                    if cell.value == test_case_name:
                        row_number = cell.row
                        break
                if row_number is not None:
                    break
            # print(f"row_number ={row_number}")

            # The data retrieved is stored in the row_data dictionary.
            if row_number is not None:
                for col_num in range(1, worksheet.max_column + 1):
                    heading = headings[col_num - 1]
                    value = worksheet.cell(row_number, col_num).value
                    row_data[heading] = value

                # When you call get_data with a specific column_name, it fetches the value associated with that
                # column name in the row corresponding to the specified test case name.
                if column_name in row_data:
                    value = row_data[column_name]
                    print(f"{column_name} = {value}")

                    return value
                else:
                    print(f"Column name '{column_name}' is not in the excel file.")
                    return None
            else:
                print(f"Test case name '{test_case_name}' not found in the excel file.")
                return None

        except Exception as e:
            print(f"Error while getting data for test case '{test_case_name}' and column '{column_name}': {str(e)}")
            return None

    @staticmethod
    def get_ddt():
        # Specify the Excel file path
        excel_file_path = TestData.EXCEL_FILE_PATH
        sheet_name = TestData.SHEET_NAME

        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            worksheet = workbook[sheet_name]

            # Retrieve the column names (headings) from the first row of the Excel sheet
            # The headings are stored in the 'headings' list.
            headings = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]

            test_data = []

            # row_data dictionary
            for row_number in range(2, worksheet.max_row + 1):
                row_data = {}
                for col_num in range(1, worksheet.max_column + 1):
                    heading = headings[col_num - 1]
                    value = worksheet.cell(row_number, col_num).value
                    row_data[heading] = value

                # appends the row_data dictionary to the test_data list.
                test_data.append(row_data)

            # test_data list, which contains dictionaries for each row of data, with keys representing column names
            # and values representing the data in those columns.
            return test_data

        except Exception as e:
            print(f"Error while getting test data: {str(e)}")
            return []
