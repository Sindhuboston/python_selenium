import openpyxl

def get_data():
    # Specify the Excel file path
    excel_file_path = "C:\\Users\\sindh\\PycharmProjects\\POMFramework_POC\\TestData\\TDexcel.xlsx"
    sheet_name = "Sheet1"
    row_number = 4  # Use an integer, not a string

    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        workSheet = workbook[sheet_name]
        headings = [workSheet.cell(1, col).value for col in range(1, workSheet.max_column + 1)]
        row_data = {}

        print("maximum col: " + str(workSheet.max_column))

        for col_num in range(1, workSheet.max_column + 1):
            heading = headings[col_num - 1]
            value = workSheet.cell(row_number, col_num).value
            row_data[heading] = value

        # Print the data
        for heading, value in row_data.items():
            print(f"{heading}: {value}")

    except Exception as e:
        print("E: ", e)
        return {}

# Call the function to execute the code
get_data()
